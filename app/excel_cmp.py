import openpyxl

def comparar_columnas_categoria(ruta_excel_a, ruta_excel_b):
    # Cargar los archivos de Excel
    wb_a = openpyxl.load_workbook(ruta_excel_a)
    ws_a = wb_a.active  # Obtener la hoja activa del Excel A
    
    wb_b = openpyxl.load_workbook(ruta_excel_b)
    ws_b = wb_b.active  # Obtener la hoja activa del Excel B
    
    # Detectar encabezados en la fila 1
    encabezados_a = {cell.value: idx for idx, cell in enumerate(ws_a[1]) if cell.value is not None}
    encabezados_b = {cell.value: idx for idx, cell in enumerate(ws_b[1]) if cell.value is not None}
    
    # Verificar que la columna "Categoría" exista en ambos archivos
    if "Categoría" not in encabezados_a or "Categoría" not in encabezados_b:
        raise KeyError(f"La columna 'Categoría' no se encontró en uno o ambos archivos.")
    
    # Índices de la columna "Categoría" en ambos archivos
    idx_categoria_a = encabezados_a["Categoría"]
    idx_categoria_b = encabezados_b["Categoría"]
    
    # Asegurarse de que la columna "Validación" exista en B; si no, crearla
    if "Validación" not in encabezados_b:
        col_validacion_idx = len(encabezados_b) + 1  # Nueva columna al final
        ws_b.cell(row=1, column=col_validacion_idx).value = "Validación"  # Agregar encabezado
    else:
        col_validacion_idx = encabezados_b["Validación"]
    
    # Leer los valores de la columna "Categoría" en ambos archivos y comparar
    for fila_idx, (fila_a, fila_b) in enumerate(zip(ws_a.iter_rows(min_row=2, values_only=True),
                                                     ws_b.iter_rows(min_row=2, values_only=True)), start=2):
        categoria_a = fila_a[idx_categoria_a]
        categoria_b = fila_b[idx_categoria_b]
        
        # Si alguno de los valores es vacío, omitir esta fila
        if not categoria_a or not categoria_b:
            continue
        
        # Comparar y escribir "ok" o "no" en la columna "Validación" del Excel B
        if categoria_a == categoria_b:
            ws_b.cell(row=fila_idx, column=col_validacion_idx + 1).value = "ok"
        else:
            ws_b.cell(row=fila_idx, column=col_validacion_idx + 1).value = "no"
    
    # Guardar los cambios en el archivo B
    wb_b.save(ruta_excel_b)
    print(f"Comparación completada. Resultados escritos en la columna 'Validación' del archivo B.")

# Rutas de los archivos Excel
ruta_a = "D:\\Python\\agents\\app\\Casos categorizados.xlsx"
ruta_b = "D:\\Python\\agents\\app\\Casos_resultados_graph_4o.xlsx"

# Llamar a la función
comparar_columnas_categoria(ruta_a, ruta_b)
