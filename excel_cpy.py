import openpyxl

def copiar_datos_excel(ruta_excel_a, ruta_excel_b):
    # Cargar los archivos de Excel
    wb_a = openpyxl.load_workbook(ruta_excel_a)
    ws_a = wb_a.active  # Obtener la hoja activa del Excel A
    
    wb_b = openpyxl.load_workbook(ruta_excel_b)
    ws_b = wb_b.active  # Obtener la hoja activa del Excel B
    
    columnas_a = ['Asunto', 'Cuerpo', 'Categoría']  # Columnas a copiar
    
    # Detectar encabezados en la fila 1 del Excel A
    encabezados_encontrados = {cell.value: idx for idx, cell in enumerate(ws_a[1]) if cell.value is not None}
    
    # Verificar que todas las columnas necesarias están presentes
    for columna in columnas_a:
        if columna not in encabezados_encontrados:
            raise KeyError(f"La columna '{columna}' no se encontró en los encabezados del Excel A. Encabezados encontrados: {list(encabezados_encontrados.keys())}")
    
    # Mapear nombres de columnas a índices
    encabezados_a = {col: encabezados_encontrados[col] for col in columnas_a}
    
    # Encontrar la primera fila vacía en el Excel B
    fila_inicio_b = ws_b.max_row + 1
    
    # Copiar datos desde la fila 2 en A hacia la primera fila vacía en B (omitimos encabezados)
    for fila in ws_a.iter_rows(min_row=2, values_only=True):  # Desde la fila 2 para saltar encabezados
        if all(dato is None for dato in fila):  # Ignorar filas completamente vacías
            continue
        datos = [fila[encabezados_a[col]] for col in columnas_a]  # Obtener valores de las columnas relevantes
        for col_idx, dato in enumerate(datos, start=1):  # Pegar en columnas 1, 2, y 3 en B
            ws_b.cell(row=fila_inicio_b, column=col_idx).value = dato
        fila_inicio_b += 1  # Avanzar a la siguiente fila en B

    # Guardar los cambios en el Excel B
    wb_b.save(ruta_excel_b)
    print(f"Datos copiados correctamente de {ruta_excel_a} a {ruta_excel_b}")

# Rutas de los archivos Excel
ruta_a = "D:\\Python\\agents\\app\\Casos_A.xlsx"
ruta_b = "D:\\Python\\agents\\app\\Casos categorizados.xlsx"

# Llamar a la función
copiar_datos_excel(ruta_a, ruta_b)
